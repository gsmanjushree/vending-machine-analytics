import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, Alignment

# Load the dataset
df = pd.read_csv('vending_sales.csv')

# Convert timestamp to datetime
df['timestamp'] = pd.to_datetime(df['timestamp'])
df['date'] = df['timestamp'].dt.date
df['revenue'] = df['price'] * df['quantity']

# Create Excel workbook
wb = Workbook()

# Remove default sheet
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Sheet 1: Raw Data
ws_data = wb.create_sheet(title="Raw Data")
for r_idx, row in enumerate(df.values, 1):
    for c_idx, value in enumerate(row, 1):
        ws_data.cell(row=r_idx, column=c_idx, value=value)

# Add headers
for c_idx, col_name in enumerate(df.columns, 1):
    ws_data.cell(row=1, column=c_idx, value=col_name)
    ws_data.cell(row=1, column=c_idx).font = Font(bold=True)

# Sheet 2: Pivot Tables
ws_pivot = wb.create_sheet(title="Pivot Analysis")

# Pivot 1: Total Sales by Product & Machine
pivot1_data = df.pivot_table(values='quantity', index='product', columns='machine_id', aggfunc='sum', fill_value=0)
ws_pivot.cell(row=1, column=1, value="Sales Volume by Product & Machine").font = Font(bold=True, size=14)

# Write pivot table 1
start_row = 3
start_col = 1

# Headers
ws_pivot.cell(row=start_row, column=start_col, value="Product")
for idx, machine in enumerate(pivot1_data.columns):
    ws_pivot.cell(row=start_row, column=start_col + idx + 1, value=machine)
ws_pivot.cell(row=start_row, column=start_col + len(pivot1_data.columns) + 1, value="Total")

# Make headers bold
for col in range(start_col, start_col + len(pivot1_data.columns) + 2):
    ws_pivot.cell(row=start_row, column=col).font = Font(bold=True)

# Data rows
for idx, (product, row) in enumerate(pivot1_data.iterrows()):
    row_num = start_row + idx + 1
    ws_pivot.cell(row=row_num, column=start_col, value=product)
    
    total = 0
    for col_idx, value in enumerate(row.values):
        ws_pivot.cell(row=row_num, column=start_col + col_idx + 1, value=int(value))
        total += value
    
    ws_pivot.cell(row=row_num, column=start_col + len(pivot1_data.columns) + 1, value=int(total))

# Add totals row
total_row = start_row + len(pivot1_data) + 1
ws_pivot.cell(row=total_row, column=start_col, value="TOTAL").font = Font(bold=True)

for col_idx in range(len(pivot1_data.columns)):
    total_val = pivot1_data.iloc[:, col_idx].sum()
    ws_pivot.cell(row=total_row, column=start_col + col_idx + 1, value=int(total_val)).font = Font(bold=True)

grand_total = pivot1_data.sum().sum()
ws_pivot.cell(row=total_row, column=start_col + len(pivot1_data.columns) + 1, value=int(grand_total)).font = Font(bold=True)

# Pivot 2: Daily Revenue by Machine
pivot2_start_row = total_row + 3
ws_pivot.cell(row=pivot2_start_row, column=1, value="Daily Revenue by Machine").font = Font(bold=True, size=14)

daily_revenue = df.groupby(['date', 'machine_id'])['revenue'].sum().unstack(fill_value=0)

# Write pivot table 2 headers
pivot2_data_start = pivot2_start_row + 2
ws_pivot.cell(row=pivot2_data_start, column=1, value="Date")
for idx, machine in enumerate(daily_revenue.columns):
    ws_pivot.cell(row=pivot2_data_start, column=2 + idx, value=machine)

# Make headers bold
for col in range(1, 2 + len(daily_revenue.columns)):
    ws_pivot.cell(row=pivot2_data_start, column=col).font = Font(bold=True)

# Data rows - limit to first 20 days for readability
limited_daily = daily_revenue.head(20)
for idx, (date, row) in enumerate(limited_daily.iterrows()):
    row_num = pivot2_data_start + idx + 1
    ws_pivot.cell(row=row_num, column=1, value=str(date))
    
    for col_idx, value in enumerate(row.values):
        ws_pivot.cell(row=row_num, column=2 + col_idx, value=round(float(value), 2))

print("Excel pivot tables created successfully!")

# Save the workbook
wb.save('vending_pivots.xlsx')
print("Excel file 'vending_pivots.xlsx' saved!")